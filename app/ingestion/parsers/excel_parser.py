from pathlib import Path
from openpyxl import load_workbook
import msoffcrypto

from app.models.models import DocumentSection
from app.exceptions import PasswordProtectedError

def parse_xlsx(file_path: str) -> list[DocumentSection]:
    with open(file_path, "rb") as f:
        office = msoffcrypto.OfficeFile(f)

        if office.is_encrypted():
            raise PasswordProtectedError(f"{file_path} is password protected")
    wb = load_workbook(
        file_path,
        data_only=True
    )

    sections = []

    for sheet_name in wb.sheetnames:

        ws = wb[sheet_name]

        rows = []

        for row in ws.iter_rows(values_only=True):

            values = [
                str(cell)
                for cell in row
                if cell is not None
            ]

            if values:
                rows.append(" | ".join(values))

        if not rows:
            continue

        sections.append(
            DocumentSection(
                text="\n".join(rows),
                source_file=Path(file_path).name,
                document_type="xlsx",
                metadata={
                    "sheet": sheet_name
                }
            )
        )

    return sections